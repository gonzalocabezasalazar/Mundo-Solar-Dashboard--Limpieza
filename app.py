import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import io
import base64
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIGURACIÓN DE PÁGINA
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Dashboard Limpieza",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─────────────────────────────────────────────
# ESTILOS CSS
# ─────────────────────────────────────────────
st.markdown("""
<style>
    /* Fondo general */
    .stApp {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    }
    
    /* Contenido principal */
    .block-container {
        background: transparent;
        padding-top: 2rem;
    }
    
    /* Título principal */
    .main-header {
        background: white;
        padding: 30px;
        border-radius: 15px;
        box-shadow: 0 10px 30px rgba(0,0,0,0.2);
        text-align: center;
        margin-bottom: 25px;
    }
    .main-header h1 {
        color: #667eea;
        font-size: 2.5em;
        margin-bottom: 5px;
    }
    .main-header p {
        color: #666;
        font-size: 1.1em;
    }
    
    /* Tarjetas KPI */
    .kpi-card {
        background: white;
        padding: 25px 20px;
        border-radius: 15px;
        box-shadow: 0 10px 30px rgba(0,0,0,0.2);
        text-align: center;
        transition: all 0.3s;
        height: 140px;
        display: flex;
        flex-direction: column;
        justify-content: center;
    }
    .kpi-label {
        color: #888;
        font-size: 0.8em;
        text-transform: uppercase;
        letter-spacing: 1px;
        margin-bottom: 10px;
        font-weight: 600;
    }
    .kpi-value {
        color: #667eea;
        font-size: 2.2em;
        font-weight: bold;
        line-height: 1;
        margin-bottom: 5px;
    }
    .kpi-sub {
        color: #bbb;
        font-size: 0.8em;
    }
    
    /* Tarjetas de gráficos */
    .chart-card {
        background: white;
        padding: 20px;
        border-radius: 15px;
        box-shadow: 0 10px 30px rgba(0,0,0,0.2);
        margin-bottom: 20px;
    }
    
    /* Sidebar */
    [data-testid="stSidebar"] {
        background: white;
        box-shadow: 5px 0 20px rgba(0,0,0,0.1);
    }
    
    /* Upload box */
    .upload-section {
        background: white;
        padding: 30px;
        border-radius: 15px;
        box-shadow: 0 10px 30px rgba(0,0,0,0.2);
        text-align: center;
        margin-bottom: 25px;
    }
    
    /* Ocultar elementos de Streamlit */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# FUNCIONES DE PROCESAMIENTO
# ─────────────────────────────────────────────

def get_tracker_column(df):
    """Detecta si la columna de tracker se llama 'CBOX' o 'Tracker'"""
    if 'Tracker' in df.columns:
        return 'Tracker'
    elif 'CBOX' in df.columns:
        return 'CBOX'
    return None

def get_strings_column(df):
    """Detecta el nombre exacto de la columna de strings"""
    for col in df.columns:
        if 'String' in col or 'string' in col:
            return col
    return None

def load_excel(file) -> dict:
    """Carga y procesa el archivo Excel"""
    try:
        xl = pd.ExcelFile(file)
        sheets = xl.sheet_names

        if 'REGISTRO_DIARIO' not in sheets:
            st.error("❌ No se encontró la hoja 'REGISTRO_DIARIO' en el archivo.")
            return None

        # Leer REGISTRO_DIARIO
        df_reg = pd.read_excel(file, sheet_name='REGISTRO_DIARIO')
        df_reg = df_reg.iloc[:, :10]

        tracker_col = get_tracker_column(df_reg)
        if not tracker_col:
            st.error("❌ No se encontró columna 'Tracker' o 'CBOX'.")
            return None

        df_reg = df_reg.dropna(subset=['Fecha', tracker_col])
        df_reg['Fecha'] = pd.to_datetime(df_reg['Fecha']).dt.date
        df_reg = df_reg.rename(columns={tracker_col: 'Tracker'})

        # Columna strings
        strings_col = get_strings_column(df_reg)
        if strings_col and strings_col != 'Strings':
            df_reg = df_reg.rename(columns={strings_col: 'Strings'})

        # Leer BASE_DATOS si existe
        df_base = None
        if 'BASE_DATOS' in sheets:
            df_base = pd.read_excel(file, sheet_name='BASE_DATOS')

        # Calcular progreso correcto día a día
        df_progreso = calcular_progreso(df_reg)

        # Nombre de planta desde el archivo
        nombre = file.name.replace('limpieza_en_seco_', '').replace('.xlsx', '').replace('.xls', '')

        return {
            'registro': df_reg,
            'base': df_base,
            'progreso': df_progreso,
            'nombre': nombre,
            'tracker_col': 'Tracker'
        }

    except Exception as e:
        st.error(f"❌ Error al procesar el archivo: {str(e)}")
        return None


def calcular_progreso(df: pd.DataFrame) -> pd.DataFrame:
    """Calcula el progreso acumulado correcto día a día"""
    total_paneles = df['Paneles Acumulados'].max() if 'Paneles Acumulados' in df.columns else df['Paneles Limpiados'].sum()

    resumen = (
        df.groupby('Fecha')['Paneles Limpiados']
        .sum()
        .reset_index()
        .sort_values('Fecha')
    )
    resumen['Acumulado'] = resumen['Paneles Limpiados'].cumsum()
    resumen['% Avance'] = (resumen['Acumulado'] / total_paneles * 100).round(2)
    resumen.columns = ['Fecha', 'Paneles del Día', 'Paneles Acumulados', '% Avance']
    return resumen


def apply_filters(df: pd.DataFrame, fecha, inversor, cbox, tracker) -> pd.DataFrame:
    """Aplica filtros al dataframe"""
    filtered = df.copy()
    if fecha != 'Todas':
        filtered = filtered[filtered['Fecha'] == pd.to_datetime(fecha).date()]
    if inversor != 'Todos':
        filtered = filtered[filtered['Inversor'] == inversor]
    if cbox != 'Todos' and 'CBOX' in df.columns:
        filtered = filtered[filtered['CBOX'] == cbox]
    if tracker != 'Todos':
        filtered = filtered[filtered['Tracker'] == tracker]
    return filtered


# ─────────────────────────────────────────────
# COMPONENTES DE VISUALIZACIÓN
# ─────────────────────────────────────────────

def render_kpis(df: pd.DataFrame, progreso: pd.DataFrame):
    """Renderiza tarjetas KPI"""
    total_paneles = int(df['Paneles Limpiados'].sum())
    total_strings = int(df['Strings'].sum()) if 'Strings' in df.columns else 0
    max_avance = float(progreso['% Avance'].max()) if len(progreso) > 0 else 0.0
    total_potencia = float(df['Potencia DC Asociada'].sum()) if 'Potencia DC Asociada' in df.columns else 0.0

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-label">Total Paneles Limpiados</div>
            <div class="kpi-value">{total_paneles:,}</div>
            <div class="kpi-sub">Acumulado</div>
        </div>
        """, unsafe_allow_html=True)

    with col2:
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-label">Strings Limpiados</div>
            <div class="kpi-value">{total_strings:,}</div>
            <div class="kpi-sub">Total</div>
        </div>
        """, unsafe_allow_html=True)

    with col3:
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-label">% Avance Total</div>
            <div class="kpi-value">{max_avance:.1f}%</div>
            <div class="kpi-sub">Progreso Global</div>
        </div>
        """, unsafe_allow_html=True)

    with col4:
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-label">Potencia DC Total</div>
            <div class="kpi-value">{total_potencia:.0f}</div>
            <div class="kpi-sub">kW</div>
        </div>
        """, unsafe_allow_html=True)


def render_charts(df: pd.DataFrame, progreso: pd.DataFrame):
    """Renderiza los 4 gráficos y retorna las figuras para PDF"""
    
    # ── Colores de marca ──────────────────────
    COLOR_PRIMARY   = '#667eea'
    COLOR_SECONDARY = '#764ba2'
    COLOR_TEAL      = '#4ecdc4'
    COLOR_RED       = '#ff6b6b'
    PALETTE = [COLOR_PRIMARY, COLOR_SECONDARY, COLOR_TEAL, COLOR_RED,
               '#a29bfe', '#fd79a8', '#00cec9', '#fdcb6e']

    col1, col2 = st.columns(2)

    # ── Gráfico 1: Paneles por Tracker ────────
    with col1:
        st.markdown('<div class="chart-card">', unsafe_allow_html=True)
        tracker_data = df.groupby('Tracker')['Paneles Limpiados'].sum().reset_index().sort_values('Tracker')
        # Convertir a tipos nativos Python para evitar problemas con numpy.int64
        t_labels = tracker_data['Tracker'].tolist()
        t_values = [int(v) for v in tracker_data['Paneles Limpiados'].tolist()]
        fig1 = go.Figure(go.Bar(
            x=t_labels,
            y=t_values,
            marker_color=COLOR_PRIMARY,
            marker_line_color=COLOR_PRIMARY,
            hovertemplate='<b>%{x}</b><br>Paneles: %{y:,}<extra></extra>'
        ))
        fig1.update_layout(
            title='📊 Paneles Limpiados por Tracker',
            plot_bgcolor='white', paper_bgcolor='white',
            title_font_color=COLOR_PRIMARY,
            showlegend=False,
            xaxis=dict(tickangle=-45, type='category'),
            height=350,
            margin=dict(t=50, b=60)
        )
        st.plotly_chart(fig1, use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

    # ── Gráfico 2: Progreso acumulado ─────────
    with col2:
        st.markdown('<div class="chart-card">', unsafe_allow_html=True)
        # Convertir fechas a string YYYY-MM-DD y valores a float nativo
        prog_labels = [str(f) for f in progreso['Fecha'].tolist()]
        prog_values = [float(v) for v in progreso['% Avance'].tolist()]
        prog_acum   = [int(v) for v in progreso['Paneles Acumulados'].tolist()]
        prog_dia    = [int(v) for v in progreso['Paneles del Día'].tolist()]
        fig2 = go.Figure()
        fig2.add_trace(go.Scatter(
            x=prog_labels,
            y=prog_values,
            mode='lines+markers',
            fill='tozeroy',
            line=dict(color=COLOR_SECONDARY, width=3),
            marker=dict(size=10, color=COLOR_SECONDARY),
            customdata=list(zip(prog_acum, prog_dia)),
            hovertemplate=(
                '<b>%{x}</b><br>'
                'Avance: %{y:.2f}%<br>'
                'Acumulado: %{customdata[0]:,} paneles<br>'
                'Hoy: %{customdata[1]:,} paneles<extra></extra>'
            )
        ))
        fig2.update_layout(
            title='📈 Progreso Acumulado por Fecha',
            title_font_color=COLOR_PRIMARY,
            plot_bgcolor='white', paper_bgcolor='white',
            yaxis=dict(range=[0, 105], ticksuffix='%'),
            xaxis=dict(type='category'),   # ← clave: categoría, no datetime
            height=350,
            margin=dict(t=50, b=40)
        )
        st.plotly_chart(fig2, use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

    col3, col4 = st.columns(2)

    # ── Gráfico 3: Potencia por Inversor ──────
    with col3:
        st.markdown('<div class="chart-card">', unsafe_allow_html=True)
        if 'Potencia DC Asociada' in df.columns:
            pot_data = df.groupby('Inversor')['Potencia DC Asociada'].sum().reset_index()
            # Convertir a tipos nativos Python
            pot_labels = pot_data['Inversor'].tolist()
            pot_values = [float(v) for v in pot_data['Potencia DC Asociada'].tolist()]
            fig3 = go.Figure(go.Pie(
                labels=pot_labels,
                values=pot_values,
                hole=0.4,
                marker=dict(colors=PALETTE[:len(pot_labels)]),
                textinfo='percent+label',
                textposition='inside',
                hovertemplate='<b>%{label}</b><br>%{value:.1f} kW<br>%{percent}<extra></extra>'
            ))
            fig3.update_layout(
                title='⚡ Potencia DC por Inversor',
                title_font_color=COLOR_PRIMARY,
                paper_bgcolor='white',
                height=350,
                margin=dict(t=50, b=40),
                legend=dict(orientation='v', x=1, y=0.5)
            )
            st.plotly_chart(fig3, use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

    # ── Gráfico 4: Paneles por Fecha ──────────
    with col4:
        st.markdown('<div class="chart-card">', unsafe_allow_html=True)
        # Convertir fechas a string para evitar interpretación como datetime
        fecha_labels = [str(f) for f in progreso['Fecha']]
        paneles_vals = progreso['Paneles del Día'].tolist()
        fig4 = go.Figure(go.Bar(
            x=fecha_labels,
            y=paneles_vals,
            text=paneles_vals,
            texttemplate='%{text:,}',
            textposition='outside',
            marker_color=COLOR_TEAL,
            marker_line_color=COLOR_TEAL,
            marker_line_width=2,
        ))
        fig4.update_layout(
            title='🎯 Paneles Limpiados por Fecha',
            title_font_color=COLOR_PRIMARY,
            plot_bgcolor='white', paper_bgcolor='white',
            showlegend=False,
            height=350,
            margin=dict(t=50, b=40),
            xaxis=dict(title='Fecha', type='category'),
            yaxis=dict(
                title='Paneles',
                range=[0, max(paneles_vals) * 1.2]
            )
        )
        st.plotly_chart(fig4, use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

    return fig1, fig2, fig3, fig4


def render_table(df: pd.DataFrame, base: pd.DataFrame):
    """Renderiza la tabla de detalle"""
    st.markdown('<div class="chart-card">', unsafe_allow_html=True)
    st.markdown("### 📋 Detalle de Registros")

    # Preparar tabla de display
    display_cols = ['Fecha', 'Tracker', 'Inversor', 'Paneles Limpiados']
    if 'Strings' in df.columns:
        display_cols.append('Strings')
    if '% Avance' in df.columns:
        display_cols.append('% Avance')
    if 'Potencia DC Asociada' in df.columns:
        display_cols.append('Potencia DC Asociada')

    df_display = df[display_cols].copy()
    df_display['Fecha'] = df_display['Fecha'].astype(str)

    if '% Avance' in df_display.columns:
        df_display['% Avance'] = (df_display['% Avance'] * 100).map('{:.0f}%'.format)

    if 'Potencia DC Asociada' in df_display.columns:
        df_display['Potencia DC Asociada'] = df_display['Potencia DC Asociada'].map('{:.1f} kW'.format)

    st.dataframe(
        df_display,
        use_container_width=True,
        hide_index=True,
        height=400,
        column_config={
            'Paneles Limpiados': st.column_config.NumberColumn(format='%d'),
        }
    )
    st.markdown('</div>', unsafe_allow_html=True)


# ─────────────────────────────────────────────
# FUNCIONES DE DESCARGA
# ─────────────────────────────────────────────

def generar_excel(df: pd.DataFrame, progreso: pd.DataFrame, planta: str) -> bytes:
    """Genera un Excel formateado con múltiples hojas"""
    wb = Workbook()

    # ── Colores ───────────────────────────────
    PURPLE      = '667EEA'
    PURPLE_DARK = '764BA2'
    TEAL        = '4ECDC4'
    LIGHT_BG    = 'F0F2FF'
    WHITE       = 'FFFFFF'
    GRAY_LIGHT  = 'F8F9FA'
    GRAY_BORDER = 'DEE2E6'

    header_font  = Font(name='Segoe UI', bold=True, color=WHITE, size=11)
    title_font   = Font(name='Segoe UI', bold=True, color=PURPLE_DARK, size=14)
    normal_font  = Font(name='Segoe UI', size=10)
    bold_font    = Font(name='Segoe UI', bold=True, size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align   = Alignment(horizontal='left',   vertical='center')

    def thin_border():
        s = Side(style='thin', color=GRAY_BORDER)
        return Border(left=s, right=s, top=s, bottom=s)

    def set_header_row(ws, row, cols, labels):
        for i, label in enumerate(labels, 1):
            c = ws.cell(row=row, column=i, value=label)
            c.fill      = PatternFill('solid', fgColor=PURPLE)
            c.font      = header_font
            c.alignment = center_align
            c.border    = thin_border()

    # ══════════════════════════════════════════
    # HOJA 1 – RESUMEN EJECUTIVO
    # ══════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Resumen Ejecutivo'
    ws1.sheet_view.showGridLines = False

    # Título
    ws1.merge_cells('A1:F1')
    t = ws1['A1']
    t.value     = f'INFORME DE LIMPIEZA — PLANTA {planta.upper()}'
    t.font      = Font(name='Segoe UI', bold=True, color=WHITE, size=16)
    t.fill      = PatternFill('solid', fgColor=PURPLE_DARK)
    t.alignment = center_align
    ws1.row_dimensions[1].height = 40

    # Fecha de emisión
    ws1.merge_cells('A2:F2')
    d = ws1['A2']
    d.value     = f'Fecha de emisión: {date.today().strftime("%d/%m/%Y")}'
    d.font      = Font(name='Segoe UI', italic=True, color='888888', size=10)
    d.fill      = PatternFill('solid', fgColor=LIGHT_BG)
    d.alignment = center_align
    ws1.row_dimensions[2].height = 20

    ws1.row_dimensions[3].height = 15  # spacer

    # ── KPI Cards ─────────────────────────────
    total_paneles  = int(df['Paneles Limpiados'].sum())
    total_strings  = int(df['Strings'].sum()) if 'Strings' in df.columns else 0
    max_avance     = float(progreso['% Avance'].max()) if len(progreso) > 0 else 0.0
    total_potencia = float(df['Potencia DC Asociada'].sum()) if 'Potencia DC Asociada' in df.columns else 0.0

    kpis = [
        ('Total Paneles\nLimpiados', f'{total_paneles:,}', PURPLE),
        ('Strings\nLimpiados',       f'{total_strings:,}', PURPLE_DARK),
        ('% Avance\nTotal',          f'{max_avance:.1f}%', '4ECDC4'),
        ('Potencia DC\nTotal',       f'{total_potencia:.0f} kW', 'FF6B6B'),
    ]

    kpi_cols = [1, 2, 3, 4]
    for row_kpi in [4, 5, 6]:
        ws1.row_dimensions[row_kpi].height = 22

    for idx, (label, value, color) in enumerate(kpis, 1):
        # Etiqueta
        lc = ws1.cell(row=4, column=idx, value=label)
        lc.fill      = PatternFill('solid', fgColor=color)
        lc.font      = Font(name='Segoe UI', bold=True, color=WHITE, size=10)
        lc.alignment = center_align
        lc.border    = thin_border()
        ws1.row_dimensions[4].height = 30

        # Valor
        vc = ws1.cell(row=5, column=idx, value=value)
        vc.fill      = PatternFill('solid', fgColor=LIGHT_BG)
        vc.font      = Font(name='Segoe UI', bold=True, color=color, size=16)
        vc.alignment = center_align
        vc.border    = thin_border()
        ws1.row_dimensions[5].height = 35

    ws1.row_dimensions[6].height = 15  # spacer

    # ── Tabla de Progreso por Fecha ───────────
    ws1['A7'].value     = 'Progreso Diario'
    ws1['A7'].font      = title_font
    ws1['A7'].alignment = left_align
    ws1.row_dimensions[7].height = 25

    headers_prog = ['Fecha', 'Paneles del Día', 'Paneles Acumulados', '% Avance']
    set_header_row(ws1, 8, range(1, 5), headers_prog)
    ws1.row_dimensions[8].height = 25

    for i, (_, row) in enumerate(progreso.iterrows(), 9):
        ws1.row_dimensions[i].height = 20
        vals = [str(row['Fecha']), row['Paneles del Día'],
                row['Paneles Acumulados'], f"{row['% Avance']:.2f}%"]
        bg = GRAY_LIGHT if i % 2 == 0 else WHITE
        for j, val in enumerate(vals, 1):
            c = ws1.cell(row=i, column=j, value=val)
            c.fill      = PatternFill('solid', fgColor=bg)
            c.font      = normal_font
            c.alignment = center_align
            c.border    = thin_border()

    # Anchos columnas
    for col, w in zip(['A','B','C','D','E','F'], [30,20,20,20,20,20]):
        ws1.column_dimensions[col].width = w

    # ══════════════════════════════════════════
    # HOJA 2 – DETALLE DE REGISTROS
    # ══════════════════════════════════════════
    ws2 = wb.create_sheet('Detalle de Registros')
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells('A1:H1')
    t2 = ws2['A1']
    t2.value     = f'Detalle de Registros — Planta {planta}'
    t2.font      = Font(name='Segoe UI', bold=True, color=WHITE, size=14)
    t2.fill      = PatternFill('solid', fgColor=PURPLE)
    t2.alignment = center_align
    ws2.row_dimensions[1].height = 35

    det_cols  = ['Fecha', 'Tracker', 'Inversor', 'Paneles Limpiados',
                 'Strings', '% Avance', 'Potencia DC Asociada']
    det_labels = ['Fecha', 'Tracker', 'Inversor', 'Paneles Limpiados',
                  'Strings', '% Avance', 'Potencia DC (kW)']

    available = [c for c in det_cols if c in df.columns]
    labels_ok = [det_labels[det_cols.index(c)] for c in available]

    set_header_row(ws2, 2, range(1, len(available)+1), labels_ok)
    ws2.row_dimensions[2].height = 25

    for i, (_, row) in enumerate(df[available].iterrows(), 3):
        ws2.row_dimensions[i].height = 18
        bg = GRAY_LIGHT if i % 2 == 0 else WHITE
        for j, col in enumerate(available, 1):
            val = row[col]
            if col == 'Fecha':
                val = str(val)
            elif col == '% Avance':
                val = f"{float(val)*100:.0f}%"
            elif col == 'Potencia DC Asociada':
                val = f"{float(val):.1f}"
            c = ws2.cell(row=i, column=j, value=val)
            c.fill      = PatternFill('solid', fgColor=bg)
            c.font      = normal_font
            c.alignment = center_align
            c.border    = thin_border()

    col_widths = [15, 15, 15, 20, 12, 12, 18]
    for idx, w in enumerate(col_widths[:len(available)], 1):
        ws2.column_dimensions[get_column_letter(idx)].width = w

    # ══════════════════════════════════════════
    # HOJA 3 – PROGRESO DETALLADO
    # ══════════════════════════════════════════
    ws3 = wb.create_sheet('Progreso por Inversor')
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells('A1:E1')
    t3 = ws3['A1']
    t3.value     = f'Resumen por Inversor — Planta {planta}'
    t3.font      = Font(name='Segoe UI', bold=True, color=WHITE, size=14)
    t3.fill      = PatternFill('solid', fgColor=PURPLE_DARK)
    t3.alignment = center_align
    ws3.row_dimensions[1].height = 35

    inv_grp = df.groupby('Inversor').agg(
        Trackers    = ('Tracker', 'nunique'),
        Paneles     = ('Paneles Limpiados', 'sum'),
        Strings     = ('Strings', 'sum') if 'Strings' in df.columns else ('Paneles Limpiados', 'count'),
        Potencia_kW = ('Potencia DC Asociada', 'sum') if 'Potencia DC Asociada' in df.columns else ('Paneles Limpiados', 'count')
    ).reset_index()

    inv_headers = ['Inversor', 'Trackers', 'Paneles Limpiados', 'Strings', 'Potencia DC (kW)']
    set_header_row(ws3, 2, range(1, 6), inv_headers)
    ws3.row_dimensions[2].height = 25

    colors_inv = [PURPLE, PURPLE_DARK, '4ECDC4', 'FF6B6B', 'A29BFE']
    for i, (_, row) in enumerate(inv_grp.iterrows(), 3):
        ws3.row_dimensions[i].height = 22
        color = colors_inv[i % len(colors_inv)]
        vals = [row['Inversor'], int(row['Trackers']),
                int(row['Paneles']), int(row['Strings']),
                f"{row['Potencia_kW']:.1f}"]
        for j, val in enumerate(vals, 1):
            c = ws3.cell(row=i, column=j, value=val)
            c.fill      = PatternFill('solid', fgColor=LIGHT_BG)
            c.font      = Font(name='Segoe UI', size=11,
                               bold=(j == 1), color=color if j == 1 else '333333')
            c.alignment = center_align
            c.border    = thin_border()

    for col, w in zip(['A','B','C','D','E'], [20,15,22,15,20]):
        ws3.column_dimensions[col].width = w

    # Guardar en buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generar_pdf_html(df: pd.DataFrame, progreso: pd.DataFrame,
                     planta: str,
                     fig_trackers, fig_progreso,
                     fig_potencia, fig_fecha) -> str:
    """Genera HTML con gráficos Plotly embebidos — funciona sin kaleido"""

    total_paneles  = int(df['Paneles Limpiados'].sum())
    total_strings  = int(df['Strings'].sum()) if 'Strings' in df.columns else 0
    max_avance     = float(progreso['% Avance'].max()) if len(progreso) > 0 else 0.0
    total_potencia = float(df['Potencia DC Asociada'].sum()) if 'Potencia DC Asociada' in df.columns else 0.0

    # Convertir cada figura a HTML div embebible (sin kaleido, solo JS)
    def fig_to_div(fig, height=300):
        fig_copy = fig
        fig_copy.update_layout(
            height=height,
            margin=dict(t=40, b=30, l=30, r=30),
            paper_bgcolor='white',
            plot_bgcolor='white',
        )
        return fig_copy.to_html(
            full_html=False,
            include_plotlyjs=False,   # se carga una sola vez abajo
            config={'displayModeBar': False}
        )

    div1 = fig_to_div(fig_trackers)
    div2 = fig_to_div(fig_progreso)
    div3 = fig_to_div(fig_potencia)
    div4 = fig_to_div(fig_fecha)

    # Filas de tabla
    table_rows = ''
    for i, (_, r) in enumerate(df.iterrows()):
        bg = '#f8f9ff' if i % 2 == 0 else 'white'
        avance   = f"{float(r.get('% Avance', 0))*100:.0f}%"
        potencia = f"{float(r.get('Potencia DC Asociada', 0)):.1f}"
        strings  = int(r['Strings']) if 'Strings' in df.columns else '-'
        table_rows += f"""
        <tr style="background:{bg};">
            <td>{r['Fecha']}</td>
            <td>{r['Tracker']}</td>
            <td>{r['Inversor']}</td>
            <td>{int(r['Paneles Limpiados']):,}</td>
            <td>{strings}</td>
            <td>{avance}</td>
            <td>{potencia} kW</td>
        </tr>"""

    # Filas de progreso
    prog_rows = ''.join(
        f"<tr><td>{r['Fecha']}</td>"
        f"<td>{int(r['Paneles del Día']):,}</td>"
        f"<td>{int(r['Paneles Acumulados']):,}</td>"
        f"<td>{r['% Avance']:.2f}%</td></tr>"
        for _, r in progreso.iterrows()
    )

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Informe Limpieza — Planta {planta}</title>
<!-- Plotly JS embebido desde CDN -->
<script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>
<style>
    * {{ margin:0; padding:0; box-sizing:border-box; }}
    body {{ font-family:'Segoe UI',Arial,sans-serif; color:#333; background:#f4f6fb; }}

    .header {{
        background: linear-gradient(135deg,#667eea,#764ba2);
        color: white;
        padding: 28px 35px;
        border-radius: 12px;
        margin: 20px;
        text-align: center;
    }}
    .header h1 {{ font-size:26px; margin-bottom:6px; }}
    .header p  {{ font-size:13px; opacity:.85; }}

    .kpis {{
        display: grid;
        grid-template-columns: repeat(4,1fr);
        gap: 15px;
        margin: 0 20px 20px;
    }}
    .kpi {{
        background: white;
        border-left: 5px solid #667eea;
        border-radius: 10px;
        padding: 18px 14px;
        text-align: center;
        box-shadow: 0 4px 12px rgba(102,126,234,.15);
    }}
    .kpi .label {{
        color: #888;
        font-size: 10px;
        text-transform: uppercase;
        letter-spacing: 1px;
        margin-bottom: 8px;
        font-weight: 600;
    }}
    .kpi .value {{
        color: #667eea;
        font-size: 24px;
        font-weight: bold;
    }}

    .section-title {{
        color: #667eea;
        font-size: 15px;
        font-weight: bold;
        margin: 25px 20px 10px;
        padding-bottom: 6px;
        border-bottom: 2px solid #667eea;
    }}

    .charts-grid {{
        display: grid;
        grid-template-columns: 1fr 1fr;
        gap: 15px;
        margin: 0 20px 20px;
    }}
    .chart-box {{
        background: white;
        border-radius: 10px;
        padding: 15px;
        box-shadow: 0 4px 12px rgba(0,0,0,.08);
    }}

    .table-wrap {{ margin: 0 20px 20px; }}
    table {{ width:100%; border-collapse:collapse; font-size:11px; }}
    thead tr {{ background:#667eea; color:white; }}
    th {{ padding:9px 8px; text-align:left; font-weight:600; font-size:10px; text-transform:uppercase; }}
    td {{ padding:7px 8px; border-bottom:1px solid #f0f0f0; }}

    .prog-table {{ width:55%; margin:0 20px 20px; }}

    .footer {{
        margin: 20px;
        text-align: center;
        color: #aaa;
        font-size: 10px;
        border-top: 1px solid #ddd;
        padding-top: 12px;
    }}

    .print-btn {{
        display: block;
        margin: 15px auto;
        padding: 12px 35px;
        background: linear-gradient(135deg,#667eea,#764ba2);
        color: white;
        border: none;
        border-radius: 8px;
        font-size: 15px;
        font-weight: 600;
        cursor: pointer;
        box-shadow: 0 4px 15px rgba(102,126,234,.4);
    }}
    @media print {{
        .print-btn {{ display:none; }}
        body {{ background:white; }}
        .chart-box {{ box-shadow:none; border:1px solid #eee; }}
        .charts-grid {{ page-break-inside:avoid; }}
    }}
</style>
</head>
<body>

<button class="print-btn" onclick="window.print()">🖨️ Imprimir / Guardar como PDF</button>

<div class="header">
    <h1>Informe de Limpieza en Seco</h1>
    <p>Planta {planta} &nbsp;·&nbsp; Generado el {date.today().strftime('%d/%m/%Y')}</p>
</div>

<div class="kpis">
    <div class="kpi">
        <div class="label">Paneles Limpiados</div>
        <div class="value">{total_paneles:,}</div>
    </div>
    <div class="kpi">
        <div class="label">Strings Limpiados</div>
        <div class="value">{total_strings:,}</div>
    </div>
    <div class="kpi">
        <div class="label">% Avance Total</div>
        <div class="value">{max_avance:.1f}%</div>
    </div>
    <div class="kpi">
        <div class="label">Potencia DC Total</div>
        <div class="value">{total_potencia:.0f} kW</div>
    </div>
</div>

<div class="section-title">Progreso Diario</div>
<table class="prog-table">
    <thead><tr>
        <th>Fecha</th><th>Paneles del Día</th>
        <th>Paneles Acumulados</th><th>% Avance</th>
    </tr></thead>
    <tbody>{prog_rows}</tbody>
</table>

<div class="section-title">Gráficos</div>
<div class="charts-grid">
    <div class="chart-box">{div1}</div>
    <div class="chart-box">{div2}</div>
    <div class="chart-box">{div3}</div>
    <div class="chart-box">{div4}</div>
</div>

<div class="section-title">Detalle de Registros</div>
<div class="table-wrap">
<table>
    <thead><tr>
        <th>Fecha</th><th>Tracker</th><th>Inversor</th>
        <th>Paneles</th><th>Strings</th><th>% Avance</th><th>Potencia DC</th>
    </tr></thead>
    <tbody>{table_rows}</tbody>
</table>
</div>

<div class="footer">
    Dashboard Limpieza en Seco · Sistema Universal de Control de Operaciones
</div>

</body>
</html>"""
    return html


# ─────────────────────────────────────────────
# SIDEBAR: FILTROS Y CARGA DE ARCHIVO
# ─────────────────────────────────────────────

with st.sidebar:
    st.markdown("""
    <div style="text-align:center; padding: 20px 0 10px;">
        <span style="font-size:2em;">⚡</span>
        <h2 style="color:#667eea; margin:5px 0;">Limpieza</h2>
        <p style="color:#888; font-size:0.85em;">Dashboard Universal</p>
    </div>
    <hr style="border-color:#eee; margin-bottom:20px;">
    """, unsafe_allow_html=True)

    st.markdown("### 📁 Cargar Archivo")
    uploaded_file = st.file_uploader(
        "Selecciona el Excel de limpieza",
        type=['xlsx', 'xls'],
        help="El archivo debe tener las hojas REGISTRO_DIARIO y BASE_DATOS"
    )

    st.markdown("---")

    if uploaded_file:
        st.markdown("### 🔍 Filtros")


# ─────────────────────────────────────────────
# CONTENIDO PRINCIPAL
# ─────────────────────────────────────────────

# Header principal
st.markdown("""
<div class="main-header">
    <h1>Dashboard Limpieza</h1>
    <p>Sistema Universal de Control de Operaciones</p>
</div>
""", unsafe_allow_html=True)


# ── Sin archivo cargado → pantalla de bienvenida ──
if not uploaded_file:
    st.markdown("""
    <div class="upload-section">
        <h2 style="color:#667eea; margin-bottom:15px;">📂 Carga tu archivo Excel</h2>
        <p style="color:#888; margin-bottom:20px;">
            Usa el panel izquierdo para seleccionar tu archivo de limpieza
        </p>
        <div style="background:#f8f9ff; padding:20px; border-radius:10px; display:inline-block; text-align:left;">
            <p style="color:#555;"><strong>El archivo debe contener:</strong></p>
            <p style="color:#666;">✅ Hoja <code>REGISTRO_DIARIO</code></p>
            <p style="color:#666;">✅ Hoja <code>BASE_DATOS</code></p>
            <p style="color:#666; margin-top:10px;"><strong>Compatible con cualquier planta:</strong></p>
            <p style="color:#666;">⚡ Planta Sauce &nbsp;|&nbsp; 🌳 Planta El Roble &nbsp;|&nbsp; 🏭 Otras plantas</p>
        </div>
    </div>
    """, unsafe_allow_html=True)
    st.stop()


# ── Procesar archivo ──────────────────────────
with st.spinner("⏳ Procesando archivo..."):
    data = load_excel(uploaded_file)

if not data:
    st.stop()

df_reg  = data['registro']
df_base = data['base']
df_prog = data['progreso']
planta  = data['nombre']


# ── Filtros en sidebar ────────────────────────
with st.sidebar:
    fechas  = ['Todas'] + [str(f) for f in sorted(df_reg['Fecha'].unique())]
    inversores = ['Todos'] + sorted(df_reg['Inversor'].dropna().unique().tolist())
    trackers   = ['Todos'] + sorted(df_reg['Tracker'].dropna().unique().tolist())

    cbox_opts = ['Todos']
    if df_base is not None and 'CBOX' in df_base.columns:
        cbox_opts += sorted(df_base['CBOX'].dropna().unique().tolist())

    sel_fecha    = st.selectbox("📅 Fecha",    fechas)
    sel_inversor = st.selectbox("🔌 Inversor", inversores)
    sel_cbox     = st.selectbox("📦 CBOX",     cbox_opts)
    sel_tracker  = st.selectbox("🎯 Tracker",  trackers)

    if st.button("🔄 Resetear Filtros", use_container_width=True):
        st.rerun()


# ── Aplicar filtros ───────────────────────────
df_filtered = apply_filters(df_reg, sel_fecha, sel_inversor, sel_cbox, sel_tracker)

if len(df_filtered) == 0:
    st.warning("⚠️ No hay datos con los filtros seleccionados.")
    st.stop()

# Recalcular progreso con datos filtrados
df_prog_filtered = calcular_progreso(df_filtered)

# Título de planta
st.markdown(f"""
<div style="background:white; padding:15px 25px; border-radius:12px;
     box-shadow:0 5px 20px rgba(0,0,0,0.15); margin-bottom:20px;
     display:flex; align-items:center; justify-content:space-between;">
    <h2 style="color:#667eea; margin:0;">Planta {planta}</h2>
    <span style="color:#888; font-size:0.9em;">
        {len(df_filtered):,} registros &nbsp;|&nbsp;
        {df_filtered['Fecha'].nunique()} días &nbsp;|&nbsp;
        {df_filtered['Tracker'].nunique()} trackers
    </span>
</div>
""", unsafe_allow_html=True)

# ── KPIs ──────────────────────────────────────
render_kpis(df_filtered, df_prog_filtered)

st.markdown("<br>", unsafe_allow_html=True)

# ── Gráficos ──────────────────────────────────
figs = render_charts(df_filtered, df_prog_filtered)
fig_trackers, fig_progreso, fig_potencia, fig_fecha = figs

st.markdown("<br>", unsafe_allow_html=True)

# ── Tabla ─────────────────────────────────────
render_table(df_filtered, df_base)

st.markdown("<br>", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# SECCIÓN DE DESCARGAS
# ─────────────────────────────────────────────
st.markdown("""
<div style="background:white; padding:25px; border-radius:15px;
     box-shadow:0 10px 30px rgba(0,0,0,0.2); margin-bottom:20px;">
    <h3 style="color:#667eea; margin-bottom:5px;">📥 Descargar Informe</h3>
    <p style="color:#888; font-size:0.9em; margin-bottom:15px;">
        Exporta el dashboard con los datos filtrados actualmente
    </p>
</div>
""", unsafe_allow_html=True)

col_xl, col_pdf = st.columns(2)

# ── Botón Excel ───────────────────────────────
with col_xl:
    with st.spinner("Preparando Excel..."):
        excel_bytes = generar_excel(df_filtered, df_prog_filtered, planta)
    st.download_button(
        label="📊 Descargar Excel",
        data=excel_bytes,
        file_name=f"Informe_Limpieza_{planta}_{date.today().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        help="Descarga el informe en formato Excel con 3 hojas: Resumen, Detalle y Progreso por Inversor"
    )

# ── Botón PDF ─────────────────────────────────
with col_pdf:
    with st.spinner("Preparando PDF..."):
        html_pdf = generar_pdf_html(
            df_filtered, df_prog_filtered, planta,
            fig_trackers, fig_progreso, fig_potencia, fig_fecha
        )
        pdf_bytes = html_pdf.encode('utf-8')

    st.download_button(
        label="📄 Descargar PDF",
        data=pdf_bytes,
        file_name=f"Informe_Limpieza_{planta}_{date.today().strftime('%Y%m%d')}.html",
        mime="text/html",
        use_container_width=True,
        help="Descarga el informe como HTML. Ábrelo en el navegador y usa Ctrl+P para guardar como PDF"
    )
    st.caption("💡 Abre el archivo en el navegador → Ctrl+P → Guardar como PDF")

# ── Footer ────────────────────────────────────
st.markdown("""
<div style="text-align:center; padding:20px; color:rgba(255,255,255,0.6); font-size:0.85em;">
    Dashboard Limpieza en Seco · Sistema Universal de Control
</div>
""", unsafe_allow_html=True)
